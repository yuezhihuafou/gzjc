# -*- coding: utf-8 -*-
"""
声音转换能量和密度曲线的API转换工具（JSON-first版本）

功能：
1. 调用API将音频文件转换为能量和密度曲线
2. 默认仅保存JSON格式（包含metadata）
3. 可选保存xlsx格式（用于调试，需 --write-xlsx）
4. 支持批量处理
5. 输出按 bearing_id 分桶到 datasets/sound_api/

API文档: https://docs.apipost.net/docs/detail/5a08548640ca000?target_id=2080ba74725004&locale=zh-cn
"""
import requests
import os
import sys
import json
import re
import argparse
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from datetime import datetime
from tqdm import tqdm

# 添加项目根目录到路径
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir, os.pardir))
if PROJECT_ROOT not in sys.path:
    sys.path.append(PROJECT_ROOT)

# ============================================================================
# 默认API配置
# ============================================================================
DEFAULT_API_CONFIG = {
    'api_url': 'http://115.236.25.110:8003/hardware/device/open-api/calculate-sound',
    'headers': {
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Accept-Encoding": "gzip, deflate, br",
        "Authorization": "",
        "Connection": "keep-alive",
        "Cookie": "JSESSIONID=node0jh4sa9ocu6qb1893asti0xnsf7.node0",  # ← 最新Token (2026-01-22 15:00 更新)
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "X-Requested-With": "XMLHttpRequest"
    },
    'form_data_params': {
        'freq1': '20',
        'freq2': '20000',
        'freqCount': '3000',
        'rate': '0',
        'type': '2',
        'sampleFrq': '192000',
        'soundType': '0',
        'engP': '1.25'
    },
    'file_param_name': 'files'
}


def get_default_config():
    """获取默认API配置"""
    config = DEFAULT_API_CONFIG
    return (
        config['api_url'],
        config['headers'].copy(),
        config['form_data_params'].copy(),
        config['file_param_name']
    )


def parse_bearing_id_from_filename(filename: str) -> Optional[str]:
    """从文件名解析 bearing_id"""
    base_name = Path(filename).stem
    
    # XJTU-SY_{bearing_id}_t{t} 格式
    pattern = r'XJTU-SY_(.+?)(?:_t?\d+)?$'
    match = re.match(pattern, base_name)
    if match:
        return match.group(1)
    
    # 普通格式：{bearing_id}_t{t}
    pattern = r'(.+?)(?:_t?\d+)?$'
    match = re.match(pattern, base_name)
    if match:
        bearing_id = match.group(1)
        if bearing_id.endswith('_wav'):
            bearing_id = bearing_id[:-4]
        return bearing_id
    
    return None


def parse_t_from_filename(filename: str) -> Optional[int]:
    """从文件名解析 t"""
    base_name = Path(filename).stem
    
    # _t000123 格式
    pattern1 = r'_t(\d+)$'
    match = re.search(pattern1, base_name)
    if match:
        return int(match.group(1))
    
    # _000123 格式（6位数字）
    pattern2 = r'_(\d{6})$'
    match = re.search(pattern2, base_name)
    if match:
        return int(match.group(1))
    
    return None


def test_sound_api(audio_file_path, api_url, headers=None, file_param_name='files', 
                   form_data_params=None, timeout=60):
    """测试声音转换能量和密度曲线的API"""
    if not os.path.exists(audio_file_path):
        print(f"错误: 路径不存在: {audio_file_path}")
        return None
    
    if os.path.isdir(audio_file_path):
        print(f"错误: 输入的是目录而不是文件: {audio_file_path}")
        return None
    
    if not os.path.isfile(audio_file_path):
        print(f"错误: 不是有效的文件: {audio_file_path}")
        return None
    
    file_ext = os.path.splitext(audio_file_path)[1].lower()
    mime_types = {
        '.wav': 'audio/wav',
        '.mp3': 'audio/mpeg',
        '.m4a': 'audio/mp4',
        '.flac': 'audio/flac',
        '.ogg': 'audio/ogg',
        '.pcm': 'audio/pcm'
    }
    content_type = mime_types.get(file_ext, 'application/octet-stream')
    
    try:
        with open(audio_file_path, 'rb') as f:
            files = {
                file_param_name: (os.path.basename(audio_file_path), f, content_type)
            }
            
            default_headers = {
                'User-Agent': 'Python-requests/Sound-API-Client/1.0'
            }
            if headers:
                default_headers.update(headers)
            
            data = {}
            if form_data_params:
                data.update(form_data_params)
            
            response = requests.post(
                api_url,
                files=files,
                data=data,
                headers=default_headers,
                timeout=timeout
            )
            
            if response.status_code == 200:
                try:
                    result = response.json()
                    print(f"[成功] API调用成功: {os.path.basename(audio_file_path)}")
                    return result
                except json.JSONDecodeError:
                    print(f"[警告] 响应不是JSON格式，状态码: {response.status_code}")
                    return None
            elif response.status_code == 401:
                print(f"[认证失败] API需要认证，状态码: {response.status_code}")
                return None
            else:
                print(f"[失败] API调用失败，状态码: {response.status_code}")
                return None
                
    except requests.exceptions.Timeout:
        print(f"[超时] 请求超时（{timeout}秒）")
        return None
    except requests.exceptions.RequestException as e:
        print(f"[错误] 请求异常: {str(e)}")
        return None
    except Exception as e:
        print(f"[错误] 处理异常: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


def parse_api_response(result, verbose=True):
    """解析API响应，提取能量和密度曲线数据"""
    import numpy as np
    
    if not isinstance(result, dict):
        if verbose:
            print(f"警告: 响应格式不是字典: {type(result)}")
        return None
    
    try:
        data = None
        
        # 格式0: rc/ret/err格式
        if 'rc' in result and 'ret' in result:
            if result.get('rc') == 0 and isinstance(result.get('ret'), list) and len(result['ret']) > 0:
                ret_item = result['ret'][0]
                if 'dataJson' in ret_item:
                    try:
                        data_json_str = ret_item['dataJson']
                        data_list = json.loads(data_json_str)
                        
                        if isinstance(data_list, list) and len(data_list) > 0:
                            frequencies = []
                            db_values = []
                            density_values = []
                            
                            for item in data_list:
                                if isinstance(item, dict):
                                    freq1 = item.get('freq1', 0)
                                    freq2 = item.get('freq2', 0)
                                    frequency = (freq1 + freq2) / 2.0
                                    
                                    frequencies.append(frequency)
                                    db_values.append(item.get('db', 0))
                                    density_values.append(item.get('density', 0))
                            
                            if frequencies and db_values and density_values:
                                data = {
                                    'frequency': np.array(frequencies, dtype=np.float32),
                                    'volume': np.array(db_values, dtype=np.float32),
                                    'density': np.array(density_values, dtype=np.float32)
                                }
                    except Exception as e:
                        if verbose:
                            print(f"解析dataJson字段时出错: {e}")
        
        # 格式1: soundFrequencyDtoList格式
        if 'soundFrequencyDtoList' in result:
            dto_list = result['soundFrequencyDtoList']
            if isinstance(dto_list, list) and len(dto_list) > 0:
                db_values = []
                density_values = []
                for item in dto_list:
                    if isinstance(item, dict):
                        db_values.append(item.get('db', 0))
                        density_values.append(item.get('density', 0))
                
                if db_values and density_values:
                    n_points = len(db_values)
                    freq_start = 20
                    freq_end = 20000
                    frequency = np.linspace(freq_start, freq_end, n_points)
                    
                    data = {
                        'frequency': frequency.astype(np.float32),
                        'volume': np.array(db_values, dtype=np.float32),
                        'density': np.array(density_values, dtype=np.float32)
                    }
        
        # 格式2-4: 其他格式
        elif all(key in result for key in ['frequency', 'volume', 'density']):
            data = {
                'frequency': np.array(result['frequency'], dtype=np.float32),
                'volume': np.array(result['volume'], dtype=np.float32),
                'density': np.array(result['density'], dtype=np.float32)
            }
        elif 'data' in result and isinstance(result['data'], dict):
            data_dict = result['data']
            if all(key in data_dict for key in ['frequency', 'volume', 'density']):
                data = {
                    'frequency': np.array(data_dict['frequency'], dtype=np.float32),
                    'volume': np.array(data_dict['volume'], dtype=np.float32),
                    'density': np.array(data_dict['density'], dtype=np.float32)
                }
        elif 'result' in result and isinstance(result['result'], dict):
            result_dict = result['result']
            if all(key in result_dict for key in ['frequency', 'volume', 'density']):
                data = {
                    'frequency': np.array(result_dict['frequency'], dtype=np.float32),
                    'volume': np.array(result_dict['volume'], dtype=np.float32),
                    'density': np.array(result_dict['density'], dtype=np.float32)
                }
        
        if data is None:
            if verbose:
                print("警告: 无法识别API响应格式")
                print(f"实际响应键: {list(result.keys())}")
            return None
        
        # 验证数据有效性
        for key in ['frequency', 'volume', 'density']:
            if len(data[key]) == 0:
                if verbose:
                    print(f"警告: {key}数组为空")
                return None
        
        # 确保长度一致
        lengths = [len(data[key]) for key in ['frequency', 'volume', 'density']]
        if len(set(lengths)) > 1:
            if verbose:
                print(f"警告: 数组长度不一致")
            min_len = min(lengths)
            for key in ['frequency', 'volume', 'density']:
                data[key] = data[key][:min_len]
        
        if verbose:
            print(f"  频率点数: {len(data['frequency'])}")
            print(f"  频率范围: {data['frequency'].min():.2f} - {data['frequency'].max():.2f} Hz")
        
        return data
        
    except Exception as e:
        if verbose:
            print(f"解析响应时出错: {str(e)}")
            import traceback
            traceback.print_exc()
        return None


def save_to_json_with_metadata(
    data: Dict,
    output_file: str,
    source_path: str,
    bearing_id: Optional[str] = None,
    t: Optional[int] = None,
    api_url: Optional[str] = None,
    form_data_params: Optional[Dict] = None,
    file_param_name: Optional[str] = None
):
    """保存数据为JSON格式（包含metadata，稳定schema）"""
    os.makedirs(os.path.dirname(output_file) if os.path.dirname(output_file) else '.', exist_ok=True)
    
    # 稳定的 metadata schema
    metadata = {
        'bearing_id': str(bearing_id) if bearing_id else None,
        't': int(t) if t is not None else None,
        'source_path': str(source_path),
        'api_url': str(api_url) if api_url else None,
        'api_params': form_data_params.copy() if form_data_params else None,
        'created_at': datetime.now().isoformat()
    }
    
    # 稳定的 JSON schema
    json_data = {
        'data': {
            'frequency': data['frequency'].tolist(),
            'volume': data['volume'].tolist(),
            'density': data['density'].tolist()
        },
        'metadata': metadata
    }
    
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(json_data, f, indent=2, ensure_ascii=False)
    
    print(f"  已保存JSON: {output_file}")


def save_to_xlsx_format(data, output_file, filename_base=None):
    """保存为xlsx格式（仅调试用）"""
    try:
        import pandas as pd
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("需要安装 pandas 和 openpyxl: pip install pandas openpyxl")
    
    import numpy as np
    
    if filename_base is None:
        filename_base = os.path.splitext(os.path.basename(output_file))[0]
    
    os.makedirs(os.path.dirname(output_file) if os.path.dirname(output_file) else '.', exist_ok=True)
    
    df = pd.DataFrame({
        'frequency': data['frequency'],
        'volume': data['volume'],
        'density': data['density']
    })
    
    sheet_name = f"{filename_base}.wav"
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2, header=False)
        
        worksheet = writer.sheets[sheet_name]
        worksheet.cell(row=1, column=1, value=sheet_name)
        worksheet.cell(row=2, column=1, value='frequency')
        worksheet.cell(row=2, column=2, value='volume')
        worksheet.cell(row=2, column=3, value='density')
    
    print(f"  已保存xlsx: {output_file}")


def batch_convert(
    audio_files: List[str],
    api_url: str,
    headers: Optional[Dict] = None,
    file_param_name: str = 'files',
    form_data_params: Optional[Dict] = None,
    output_root: str = 'datasets/sound_api',
    save_json: bool = True,
    save_xlsx: bool = False
) -> Dict:
    """批量转换多个音频文件（按 bearing_id 分桶）"""
    results = {
        'success': 0,
        'failed': 0,
        'files': []
    }
    
    print(f"\n开始批量转换 {len(audio_files)} 个文件..")
    print(f"输出根目录: {output_root}")
    print(f"  JSON: {os.path.join(output_root, 'output_json', '{bearing_id}')}")
    if save_xlsx:
        print(f"  xlsx: {os.path.join(output_root, 'output_xlsx', '{bearing_id}')}")
    print()
    
    for audio_file in tqdm(audio_files, desc="转换中"):
        filename_base = os.path.splitext(os.path.basename(audio_file))[0]
        
        bearing_id = parse_bearing_id_from_filename(filename_base)
        t = parse_t_from_filename(filename_base)
        
        # 警告：无法解析 bearing_id
        if not bearing_id:
            print(f"警告: 无法从文件名解析 bearing_id: {filename_base}，使用 'unknown'")
            bearing_id = 'unknown'
        
        result = test_sound_api(
            audio_file,
            api_url,
            headers=headers,
            file_param_name=file_param_name,
            form_data_params=form_data_params
        )
        
        if result:
            data = parse_api_response(result, verbose=False)
            
            if data:
                # 按 bearing_id 分桶
                json_output_dir = os.path.join(output_root, 'output_json', str(bearing_id))
                xlsx_output_dir = os.path.join(output_root, 'output_xlsx', str(bearing_id))
                
                if save_json:
                    json_file = os.path.join(json_output_dir, f"{filename_base}.json")
                    save_to_json_with_metadata(
                        data,
                        json_file,
                        source_path=audio_file,
                        bearing_id=bearing_id,
                        t=t,
                        api_url=api_url,
                        form_data_params=form_data_params,
                        file_param_name=file_param_name
                    )
                
                if save_xlsx:
                    xlsx_file = os.path.join(xlsx_output_dir, f"{filename_base}.xlsx")
                    save_to_xlsx_format(data, xlsx_file, filename_base)
                
                results['success'] += 1
                results['files'].append({
                    'file': audio_file,
                    'status': 'success',
                    'data_points': len(data['frequency']),
                    'bearing_id': bearing_id,
                    't': t
                })
            else:
                results['failed'] += 1
                results['files'].append({
                    'file': audio_file,
                    'status': 'failed',
                    'reason': 'parse_error'
                })
        else:
            results['failed'] += 1
            results['files'].append({
                'file': audio_file,
                'status': 'failed',
                'reason': 'api_error'
            })
    
    print(f"\n转换完成: 成功 {results['success']}, 失败 {results['failed']}")
    return results


def main():
    """主函数（支持CLI和交互式模式）"""
    parser = argparse.ArgumentParser(
        description='声音转换能量和密度曲线的API转换工具（JSON-first版本）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 批量转换音频文件（默认仅保存JSON，按bearing_id分桶）
  python convert_sound_api.py --audio_dir /path/to/audio --output_root datasets/sound_api
  # 输出到: datasets/sound_api/output_json/{bearing_id}/...
  
  # 同时保存JSON和xlsx（调试用）
  python convert_sound_api.py --audio_dir /path/to/audio --write-xlsx
  # 输出到: datasets/sound_api/output_json/{bearing_id}/ 和 datasets/sound_api/output_xlsx/{bearing_id}/
  
  # 转换单个文件
  python convert_sound_api.py --audio_file /path/to/audio.wav
  
注意: 所有产物落在 datasets/sound_api/ 下，不会在 tools/sound_api/ 下落任何数据
        """
    )
    
    parser.add_argument('--audio_file', type=str, help='单个音频文件路径')
    parser.add_argument('--audio_dir', type=str, help='音频文件目录路径')
    parser.add_argument('--output_root', type=str, default='datasets/sound_api',
                       help='输出根目录（默认: datasets/sound_api）')
    parser.add_argument('--write-xlsx', action='store_true',
                       help='同时保存xlsx格式（默认仅保存JSON）')
    parser.add_argument('--api_url', type=str, help='API接口URL（可选）')
    parser.add_argument('--jsessionid', type=str, 
                       help='Tomcat JSESSIONID token（可选，如果提供则覆盖默认token）')
    parser.add_argument('--no-json', action='store_true',
                       help='不保存JSON格式（仅与--write-xlsx一起使用）')
    
    args = parser.parse_args()
    
    # 如果没有CLI参数，进入交互式模式
    if not args.audio_file and not args.audio_dir:
        print("=" * 60)
        print("声音转换能量和密度曲线API转换工具")
        print("提示: 使用 --help 查看CLI参数")
        return
    
    # CLI模式
    if args.api_url:
        api_url = args.api_url
        _, headers, form_data_params, file_param_name = get_default_config()
    else:
        api_url, headers, form_data_params, file_param_name = get_default_config()
    
    # 如果提供了JSESSIONID，更新headers中的Cookie
    if args.jsessionid:
        headers['Cookie'] = f"JSESSIONID={args.jsessionid}"
        print(f"[信息] 使用自定义JSESSIONID token")
    else:
        print(f"[信息] 使用默认JSESSIONID token (更新于2026-01-22)")
    
    save_json = not args.no_json
    save_xlsx = args.write_xlsx
    
    if args.audio_file:
        if not os.path.exists(args.audio_file):
            print(f"错误: 文件不存在: {args.audio_file}")
            return
        
        result = test_sound_api(
            args.audio_file,
            api_url,
            headers=headers,
            file_param_name=file_param_name,
            form_data_params=form_data_params
        )
        
        if result:
            data = parse_api_response(result, verbose=True)
            if data:
                filename_base = os.path.splitext(os.path.basename(args.audio_file))[0]
                bearing_id = parse_bearing_id_from_filename(filename_base)
                t = parse_t_from_filename(filename_base)
                
                if bearing_id:
                    json_output_dir = os.path.join(args.output_root, 'output_json', str(bearing_id))
                    xlsx_output_dir = os.path.join(args.output_root, 'output_xlsx', str(bearing_id))
                else:
                    json_output_dir = os.path.join(args.output_root, 'output_json', 'unknown')
                    xlsx_output_dir = os.path.join(args.output_root, 'output_xlsx', 'unknown')
                
                if save_json:
                    json_file = os.path.join(json_output_dir, f"{filename_base}.json")
                    save_to_json_with_metadata(
                        data, json_file, args.audio_file,
                        bearing_id, t, api_url, form_data_params, file_param_name
                    )
                
                if save_xlsx:
                    xlsx_file = os.path.join(xlsx_output_dir, f"{filename_base}.xlsx")
                    save_to_xlsx_format(data, xlsx_file, filename_base)
    
    elif args.audio_dir:
        if not os.path.exists(args.audio_dir):
            print(f"错误: 目录不存在: {args.audio_dir}")
            return
        
        audio_extensions = ['.wav', '.mp3', '.m4a', '.flac', '.ogg', '.pcm']
        audio_files = []
        audio_dir_path = Path(args.audio_dir)
        # 递归查找所有音频文件
        for ext in audio_extensions:
            audio_files.extend(audio_dir_path.rglob(f'*{ext}'))
            audio_files.extend(audio_dir_path.rglob(f'*{ext.upper()}'))
        
        if not audio_files:
            print(f"错误: 在 {args.audio_dir} 中未找到音频文件")
            return
        
        audio_files = [str(f) for f in audio_files]
        print(f"\n找到 {len(audio_files)} 个音频文件")
        
        results = batch_convert(
            audio_files, api_url, headers, file_param_name,
            form_data_params, args.output_root, save_json, save_xlsx
        )
        
        report_file = os.path.join(args.output_root, 'logs', 'conversion_report.json')
        os.makedirs(os.path.dirname(report_file), exist_ok=True)
        with open(report_file, 'w', encoding='utf-8') as f:
            json.dump(results, f, indent=2, ensure_ascii=False)
        print(f"\n转换报告已保存: {report_file}")


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n用户中断")
    except Exception as e:
        print(f"\n\n发生错误: {str(e)}")
        import traceback
        traceback.print_exc()
