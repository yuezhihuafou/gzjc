"""
一键从 sound_api JSON/xlsx 构建 NPZ 缓存

功能：
1. 优先从 JSON 文件读取（包含metadata），回退到 xlsx
2. bearing_id/t 解析（按优先级1→4实现）
3. 质量门禁（长度3000、finite、std、non_zero_ratio等）
4. 断点续跑（目标npz存在则跳过）
5. 并行处理（ThreadPoolExecutor/ProcessPoolExecutor）
6. 统计输出与 bad_files.txt
7. t 连续性硬校验

输入：datasets/sound_api/output_json/ 或 datasets/sound_api/output_xlsx/（递归扫描）
输出：datasets/sound_api/cache_npz/{bearing_id}/{t:06d}.npz
"""
import os
import re
import json
import argparse
import time
from pathlib import Path
from typing import List, Dict, Tuple, Optional
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, ProcessPoolExecutor, as_completed
import multiprocessing

import numpy as np
from tqdm import tqdm

from core.json_io import loadFirstJson

# 只在需要xlsx回退时导入
def _load_pandas_if_needed():
    """懒加载pandas（仅在需要xlsx回退时）"""
    try:
        import pandas as pd
        return pd
    except ImportError:
        raise ImportError("需要安装 pandas 才能读取 xlsx 文件: pip install pandas")


def _load_openpyxl_if_needed():
    """懒加载openpyxl（仅在需要xlsx回退时）"""
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        raise ImportError("需要安装 openpyxl 才能读取 xlsx 文件: pip install openpyxl")


def parse_bearing_id_from_filename(filename: str) -> Optional[str]:
    """从文件名解析 bearing_id（优先级1）"""
    base_name = Path(filename).stem
    
    # XJTU-SY_{bearing_id}_t{t} 或 XJTU-SY_{bearing_id}_{t} 格式
    pattern = r'XJTU-SY_(.+?)(?:_t?\d+)?$'
    match = re.match(pattern, base_name)
    if match:
        return match.group(1)
    
    # 普通格式：{bearing_id}_t{t} 或 {bearing_id}_{t}
    pattern = r'(.+?)(?:_t?\d+)?$'
    match = re.match(pattern, base_name)
    if match:
        bearing_id = match.group(1)
        if bearing_id.endswith('_wav'):
            bearing_id = bearing_id[:-4]
        return bearing_id
    
    return None


def parse_t_from_filename(filename: str) -> Optional[int]:
    """从文件名解析 t（优先级1）"""
    base_name = Path(filename).stem
    
    # _t000123 格式
    pattern1 = r'_t(\d+)$'
    match = re.search(pattern1, base_name)
    if match:
        return int(match.group(1))
    
    # _000123 格式（6位数字）
    pattern2 = r'_(\d{6})$'
    match = re.search(pattern2, base_name)
    if match:
        return int(match.group(1))
    
    return None


def parse_bearing_id_from_json_metadata(json_path: str) -> Optional[str]:
    """从JSON metadata解析 bearing_id（优先级2）；兼容文件末尾带时间戳等 extra 的 output_json"""
    try:
        data = loadFirstJson(json_path)
        metadata = data.get('metadata', {})
            bearing_id = metadata.get('bearing_id')
            if bearing_id:
                return str(bearing_id)
    except Exception as e:
        pass
    return None


def parse_t_from_json_metadata(json_path: str) -> Optional[int]:
    """从JSON metadata解析 t（优先级2）；兼容文件末尾带时间戳等 extra 的 output_json"""
    try:
        data = loadFirstJson(json_path)
        metadata = data.get('metadata', {})
            t = metadata.get('t')
            if t is not None:
                return int(t)
    except Exception as e:
        pass
    return None


def parse_bearing_id_from_xlsx_first_row(xlsx_path: str) -> Optional[str]:
    """从xlsx第1行解析 bearing_id（优先级3）"""
    try:
        openpyxl = _load_openpyxl_if_needed()
        from openpyxl import load_workbook
        
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        
        first_cell = ws.cell(row=1, column=1)
        if first_cell.value:
            text = str(first_cell.value)
            pattern = r'XJTU-SY_(.+)'
            match = re.search(pattern, text)
            if match:
                wb.close()
                return match.group(1)
        
        wb.close()
    except Exception as e:
        pass
    return None


def load_data_from_json(json_path: str) -> Optional[Dict]:
    """从JSON文件加载数据（优先）；兼容文件末尾带时间戳等 extra 的 output_json"""
    try:
        data = loadFirstJson(json_path)

        # 支持新格式（带metadata）
        if 'data' in data and 'metadata' in data:
            data_dict = data['data']
            metadata = data['metadata']
        else:
            # 兼容旧格式（只有三数组）
            data_dict = data
            metadata = {}

        frequency = np.array(data_dict.get('frequency', []), dtype=np.float32)
        volume = np.array(data_dict.get('volume', []), dtype=np.float32)
        density = np.array(data_dict.get('density', []), dtype=np.float32)

        return {
            'frequency': frequency,
            'volume': volume,
            'density': density,
            'metadata': metadata
        }
    except Exception as e:
        return None


def load_data_from_xlsx(xlsx_path: str) -> Optional[Dict]:
    """从xlsx文件加载数据（回退）"""
    try:
        pd = _load_pandas_if_needed()
        import pandas as pd
        
        # 跳过前2行，读取3000行三列
        df = pd.read_excel(xlsx_path, header=None, skiprows=2, nrows=3000)
        
        if df.shape[0] < 3000 or df.shape[1] < 3:
            return None
        
        frequency = df.iloc[:, 0].values.astype(np.float32)
        volume = df.iloc[:, 1].values.astype(np.float32)
        density = df.iloc[:, 2].values.astype(np.float32)
        
        return {
            'frequency': frequency,
            'volume': volume,
            'density': density,
            'metadata': {}
        }
    except Exception as e:
        return None


def quality_check(
    frequency: np.ndarray,
    volume: np.ndarray,
    density: np.ndarray,
    min_length: int = 3000,
    min_non_zero_ratio: float = 0.05
) -> Tuple[bool, str]:
    """
    质量门禁检查
    
    Returns:
        (is_valid, reason): 是否通过门禁，失败原因
    """
    # 检查长度
    if len(frequency) != min_length or len(volume) != min_length or len(density) != min_length:
        return False, f"长度不符合要求: frequency={len(frequency)}, volume={len(volume)}, density={len(density)}, 期望={min_length}"
    
    # 检查finite
    if not np.all(np.isfinite(frequency)):
        return False, "frequency 包含非有限值"
    if not np.all(np.isfinite(volume)):
        return False, "volume 包含非有限值"
    if not np.all(np.isfinite(density)):
        return False, "density 包含非有限值"
    
    # 检查 volume < 0（方案2：判坏样本）
    if np.any(volume < 0):
        return False, f"volume 包含负值（共 {np.sum(volume < 0)} 个）"
    
    # 检查 non_zero_ratio
    non_zero_ratio = np.sum(volume > 0) / len(volume)
    if non_zero_ratio < min_non_zero_ratio:
        return False, f"volume 非零比例过低: {non_zero_ratio:.4f} < {min_non_zero_ratio}"
    
    # 检查 std（基于 log1p(volume) 和 density）
    log1p_volume = np.log1p(volume)
    if np.std(log1p_volume) < 1e-6:
        return False, "log1p(volume) 标准差过小（几乎为常数）"
    if np.std(density) < 1e-6:
        return False, "density 标准差过小（几乎为常数）"
    
    return True, ""


def process_single_sample_structured(
    task: Dict,
    output_dir: str,
    skip_existing: bool = True,
    min_length: int = 3000,
    min_non_zero_ratio: float = 0.05
) -> Tuple[bool, str, Optional[str], Optional[int]]:
    """
    处理单个样本（结构化任务对象）
    
    Args:
        task: 任务对象 {'source_path', 'bearing_id', 't', 'source_type'}
        output_dir: 输出目录
        skip_existing: 是否跳过已存在的npz
        min_length: 最小长度
        min_non_zero_ratio: 最小非零比例
    
    Returns:
        (success, message, bad_reason, actual_t, fault_label): 成功标志、消息、失败原因、实际使用的t、标签（-1表示缺失）
    """
    source_path = task['source_path']
    bearing_id = task['bearing_id']
    t = task['t']
    orig_t = task.get('original_t', None)  # 原始 t（可能为 None）
    source_type = task['source_type']
    
    # 加载数据
    data = None
    if source_type == 'json':
        data = load_data_from_json(source_path)
    elif source_type == 'xlsx':
        data = load_data_from_xlsx(source_path)
    
    if data is None:
        return False, f"无法加载数据: {source_path}", "load_error", None, -1
    
    # 从 metadata 还原标签（与当前 JSON 一一对应）
    metadata = data.get('metadata', {})
    raw_label = metadata.get('fault_label', metadata.get('health_label', -1))
    if raw_label is None or (isinstance(raw_label, (int, float)) and raw_label < 0):
        fault_label = -1
    else:
        fault_label = int(raw_label)
    
    # 质量门禁
    frequency = data['frequency']
    volume = data['volume']
    density = data['density']
    
    is_valid, reason = quality_check(frequency, volume, density, min_length, min_non_zero_ratio)
    if not is_valid:
        return False, f"质量门禁失败: {reason}", f"quality_check_failed: {reason}", None, -1
    
    # 变换数据：x[0]=log1p(volume), x[1]=density
    x = np.zeros((2, min_length), dtype=np.float32)
    x[0] = np.log1p(volume)
    x[1] = density
    
    # 准备输出路径
    bearing_dir = Path(output_dir) / str(bearing_id)
    bearing_dir.mkdir(parents=True, exist_ok=True)
    
    npz_path = bearing_dir / f"{t:06d}.npz"
    
    # 断点续跑：如果已存在且skip_existing=True，则跳过（不返回 fault_label，index 中该条可不带）
    if skip_existing and npz_path.exists():
        return True, f"跳过（已存在）: {npz_path}", None, t, None
    
    # 原子写入：临时文件 + rename
    # 注意：np.savez_compressed 会自动添加 .npz 后缀，所以临时文件必须以 .npz 结尾
    temp_path = npz_path.with_suffix('.tmp.npz')
    try:
        save_dict = {
            'x': x,  # (2, 3000) float32
            'frequency': frequency,
            'bearing_id': str(bearing_id),
            't': int(t),  # 重编号后的 t（用于训练）
            'source_path': str(source_path),
            'fault_label': fault_label,  # 始终写入，-1 表示该样本无标签
        }
        
        # 保存原始 t（如果存在）
        if orig_t is not None:
            save_dict['orig_t'] = int(orig_t)
        
        np.savez_compressed(temp_path, **save_dict)
        
        # 原子重命名
        if temp_path.exists():
            temp_path.replace(npz_path)
        else:
            # 极少数情况下（如numpy版本差异），可能没有自动加后缀？保险起见检查一下
            alt_path = Path(str(temp_path) + '.npz')
            if alt_path.exists():
                alt_path.replace(npz_path)
            else:
                raise FileNotFoundError(f"临时文件未生成: {temp_path}")
        
        return True, f"成功: {npz_path}", None, t, fault_label
    except Exception as e:
        # 清理临时文件
        if temp_path.exists():
            try:
                temp_path.unlink()
            except:
                pass
        # 尝试清理可能带自动后缀的文件
        try:
            Path(str(temp_path) + '.npz').unlink(missing_ok=True)
        except:
            pass
            
        return False, f"写入失败: {str(e)}", f"write_error: {str(e)}", None, -1


def scan_input_files_recursive(
    json_dir: Optional[str] = None,
    xlsx_dir: Optional[str] = None
) -> List[Dict]:
    """
    递归扫描输入目录，收集所有 JSON 和 xlsx 文件
    
    Returns:
        List[Dict]: 每个元素包含 {'path': str, 'type': 'json'|'xlsx', 'priority': int}
    """
    files = []
    json_files_set = set()
    
    # 扫描 JSON 文件（优先）
    if json_dir and os.path.exists(json_dir):
        json_path = Path(json_dir)
        for json_file in json_path.rglob('*.json'):
            files.append({
                'path': str(json_file),
                'type': 'json',
                'priority': 1
            })
            # 记录 stem（不含扩展名），用于跳过同名xlsx
            json_files_set.add(json_file.stem)
    
    # 扫描 xlsx 文件（回退，但需要检查是否有同名JSON）
    if xlsx_dir and os.path.exists(xlsx_dir):
        xlsx_path = Path(xlsx_dir)
        for xlsx_file in xlsx_path.rglob('*.xlsx'):
            # 只有当没有同名JSON时才添加
            if xlsx_file.stem not in json_files_set:
                files.append({
                    'path': str(xlsx_file),
                    'type': 'xlsx',
                    'priority': 2
                })
    
    return files


def build_structured_tasks(files: List[Dict], t_policy: str = 'renumber') -> List[Dict]:
    """
    构建结构化任务列表（解析 bearing_id 和 t，为缺失的 t 分配值）
    
    Args:
        files: 输入文件列表
        t_policy: t 分配策略
            - 'renumber': 强制重编号，按文件名排序直接赋 t=0..T-1（推荐，保证连续）
            - 'fill_gaps': 保留显式 t，只为缺失的 t 补齐最小缺口
    
    Returns:
        List[Dict]: 每个元素包含 {'source_path', 'bearing_id', 't', 'source_type', 'original_t'}
    """
    total_files = len(files)
    print(f"  [parse] 共 {total_files} 个输入文件，开始解析 bearing_id / t ...")

    # 第一遍：解析所有文件的 bearing_id 和 t
    parsed_files: List[Dict] = []
    missing_t_initial = 0  # 记录在所有解析策略之后仍然没有显式 t 的文件数量

    for file_info in tqdm(files, desc="解析 bearing_id/t", unit="file"):
        source_path = file_info['path']
        source_type = file_info['type']
        source_path_obj = Path(source_path)
        
        # 优先级1：从文件名解析
        bearing_id = parse_bearing_id_from_filename(source_path_obj.name)
        t = parse_t_from_filename(source_path_obj.name)
        
        # 优先级2：从JSON metadata解析
        if source_type == 'json':
            if bearing_id is None:
                bearing_id = parse_bearing_id_from_json_metadata(source_path)
            if t is None:
                t = parse_t_from_json_metadata(source_path)
        # 优先级3：从xlsx第1行解析
        elif source_type == 'xlsx':
            if bearing_id is None:
                bearing_id = parse_bearing_id_from_xlsx_first_row(source_path)
        
        # 记录无法解析出的原始 t（后续会由 t_policy 统一重编号或补齐）
        if t is None:
            missing_t_initial += 1
        
        # 优先级4：如果仍无法解析 bearing_id，使用文件名基础部分
        if bearing_id is None:
            bearing_id = source_path_obj.stem
        
        parsed_files.append({
            'source_path': source_path,
            'bearing_id': bearing_id,
            't': t,  # 原始解析的 t（可能为 None）
            'original_t': t,  # 保留原始 t 用于日志/调试
            'source_type': source_type
        })

    if missing_t_initial > 0:
        print(f"  [parse] 有 {missing_t_initial} 个文件未能从文件名/metadata 中解析出原始 t，将按 t_policy='{t_policy}' 进行重编号或补齐。")
    else:
        print("  [parse] 所有文件均成功解析出原始 t 或不需要显式 t。")
    
    # 第二遍：按 bearing_id 分组，为缺失的 t 分配值
    bearing_groups = defaultdict(list)
    for parsed in parsed_files:
        bearing_groups[parsed['bearing_id']].append(parsed)
    
    # 为每个 bearing 分配 t（根据策略）
    tasks = []
    for bearing_id, group in bearing_groups.items():
        # 按文件名排序（保证可复现）
        group.sort(key=lambda x: Path(x['source_path']).name)
        
        if t_policy == 'renumber':
            # 策略A：强制重编号（忽略原 t，直接按顺序赋值 0..T-1）
            for idx, item in enumerate(group):
                item['t'] = idx
                tasks.append(item)
        
        elif t_policy == 'fill_gaps':
            # 策略B：保留显式 t，为缺失的 t 补齐最小缺口
            existing_t = [item['t'] for item in group if item['t'] is not None]
            
            if not existing_t:
                # 完全没有显式 t：赋 0..T-1
                for idx, item in enumerate(group):
                    item['t'] = idx
                    tasks.append(item)
            else:
                # 有显式 t：检查是否合法并补齐缺口
                existing_t_set = set(existing_t)
                T = len(group)
                expected_range = set(range(T))
                
                # 检查显式 t 是否在合法范围内（0..T-1）
                if not existing_t_set.issubset(expected_range):
                    # 显式 t 超出范围：回退到重编号
                    print(f"警告: bearing {bearing_id} 的显式 t 超出范围 0..{T-1}，改用重编号策略")
                    for idx, item in enumerate(group):
                        item['t'] = idx
                        tasks.append(item)
                else:
                    # 显式 t 合法：为缺失的 t 补齐
                    missing_t = sorted(expected_range - existing_t_set)
                    missing_idx = 0
                    
                    for item in group:
                        if item['t'] is None:
                            if missing_idx < len(missing_t):
                                item['t'] = missing_t[missing_idx]
                                missing_idx += 1
                            else:
                                # 不应该发生（除非逻辑错误）
                                item['t'] = T + missing_idx
                                missing_idx += 1
                        tasks.append(item)
        else:
            raise ValueError(f"未知的 t_policy: {t_policy}")
    
    return tasks


def check_original_t_quality(tasks: List[Dict], threshold: float = 0.3) -> Dict[str, Dict]:
    """
    检查原始 t 的质量（针对 renumber 模式的告警）
    
    检测逻辑：
    - 如果某 bearing 的文件有显式的 original_t（不为 None）
    - 但这些 original_t 存在"乱序/重复/跳号"问题
    - 且问题比例超过阈值 threshold，则记录警告
    
    Args:
        tasks: 任务列表（包含 bearing_id, t, original_t 等）
        threshold: 问题比例阈值（0-1），超过此值才警告
    
    Returns:
        Dict[bearing_id, Dict]: 有问题的 bearing 及其诊断信息
            {
                'bearing_id': {
                    'total_files': int,
                    'files_with_orig_t': int,
                    'issues': {
                        'disorder': int,  # 乱序数量
                        'duplicate': int,  # 重复数量
                        'gaps': int,  # 跳号数量
                    },
                    'problem_ratio': float,
                    'orig_t_values': List[int],  # 用于调试
                }
            }
    """
    # 按 bearing_id 分组
    bearing_groups = defaultdict(list)
    for task in tasks:
        bearing_groups[task['bearing_id']].append(task)
    
    warnings = {}
    
    for bearing_id, group in bearing_groups.items():
        # 只看有 original_t 的文件
        items_with_orig_t = [
            item for item in group 
            if item.get('original_t') is not None
        ]
        
        total_files = len(group)
        files_with_orig_t = len(items_with_orig_t)
        
        # 如果没有显式 original_t，跳过
        if files_with_orig_t == 0:
            continue
        
        # 提取 original_t 列表（按文件名排序，与 renumber 逻辑一致）
        items_with_orig_t.sort(key=lambda x: Path(x['source_path']).name)
        orig_t_values = [item['original_t'] for item in items_with_orig_t]
        
        # 诊断问题
        issues = {
            'disorder': 0,  # 乱序
            'duplicate': 0,  # 重复
            'gaps': 0,  # 跳号（非连续）
        }
        
        # 1. 检查重复
        if len(orig_t_values) != len(set(orig_t_values)):
            from collections import Counter
            counter = Counter(orig_t_values)
            issues['duplicate'] = sum(1 for count in counter.values() if count > 1)
        
        # 2. 检查乱序（不是严格递增）
        for i in range(1, len(orig_t_values)):
            if orig_t_values[i] <= orig_t_values[i-1]:
                issues['disorder'] += 1
        
        # 3. 检查跳号（不是连续的 0..N-1）
        sorted_orig_t = sorted(set(orig_t_values))
        if len(sorted_orig_t) > 0:
            expected_range = set(range(len(sorted_orig_t)))
            actual_range = set(sorted_orig_t)
            # 计算不匹配的值数量（缺失的 + 多出的）
            missing = expected_range - actual_range
            extra = actual_range - expected_range
            issues['gaps'] = len(missing) + len(extra)
        
        # 计算问题比例
        total_issues = issues['disorder'] + issues['duplicate'] + issues['gaps']
        problem_ratio = total_issues / files_with_orig_t if files_with_orig_t > 0 else 0
        
        # 如果问题比例超过阈值，记录警告
        if problem_ratio >= threshold:
            warnings[bearing_id] = {
                'total_files': total_files,
                'files_with_orig_t': files_with_orig_t,
                'issues': issues,
                'problem_ratio': problem_ratio,
                'orig_t_values': orig_t_values[:20],  # 只记录前20个用于调试
            }
    
    return warnings


def verify_t_continuity(output_dir: str) -> Dict[str, List[int]]:
    """
    校验每个 bearing 的 t 是否连续递增
    
    Returns:
        Dict[bearing_id, List[缺失的t]]: 不连续的 bearing 及其缺失的 t
    """
    output_path = Path(output_dir)
    if not output_path.exists():
        return {}
    
    bad_bearings = {}
    
    for bearing_dir in output_path.iterdir():
        if not bearing_dir.is_dir():
            continue
        
        bearing_id = bearing_dir.name
        
        # 收集该 bearing 下的所有 t
        t_list = []
        for npz_file in bearing_dir.glob('*.npz'):
            try:
                t = int(npz_file.stem)
                t_list.append(t)
            except ValueError:
                continue
        
        if not t_list:
            continue
        
        t_list.sort()
        T = len(t_list)
        
        # 检查是否等于 0..T-1
        expected = list(range(T))
        if t_list != expected:
            # 找出缺失的 t
            missing = set(expected) - set(t_list)
            bad_bearings[bearing_id] = sorted(missing)
    
    return bad_bearings


def main():
    parser = argparse.ArgumentParser(
        description='一键从 sound_api JSON/xlsx 构建 NPZ 缓存',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 使用默认设置（从 datasets/sound_api/output_json 构建）
  python build_sound_api_cache.py
  
  # 指定输入输出目录和工作进程数
  python build_sound_api_cache.py --json_dir datasets/sound_api/output_json --xlsx_dir datasets/sound_api/output_xlsx --output_dir datasets/sound_api/cache_npz --workers 8
  
  # 禁用断点续跑（重新处理所有文件）
  python build_sound_api_cache.py --no-skip-existing
        """
    )
    
    parser.add_argument(
        '--json_dir',
        type=str,
        default='datasets/sound_api/output_json',
        help='JSON 输入目录（递归扫描，默认: datasets/sound_api/output_json）'
    )
    parser.add_argument(
        '--xlsx_dir',
        type=str,
        default='datasets/sound_api/output_xlsx',
        help='xlsx 输入目录（递归扫描，回退用，默认: datasets/sound_api/output_xlsx）'
    )
    parser.add_argument(
        '--output_dir',
        type=str,
        default='datasets/sound_api/cache_npz',
        help='输出目录（默认: datasets/sound_api/cache_npz）'
    )
    parser.add_argument(
        '--workers',
        type=int,
        default=None,
        help='并行工作进程数（默认: min(32, CPU核心数)）'
    )
    parser.add_argument(
        '--no-skip-existing',
        action='store_true',
        help='不跳过已存在的npz文件（重新处理）'
    )
    parser.add_argument(
        '--min-length',
        type=int,
        default=3000,
        help='最小数据长度（默认: 3000）'
    )
    parser.add_argument(
        '--min-non-zero-ratio',
        type=float,
        default=0.05,
        help='最小非零比例（默认: 0.05）'
    )
    parser.add_argument(
        '--bad-files',
        type=str,
        default='datasets/sound_api/logs/bad_files_cache.txt',
        help='问题文件列表路径（默认: datasets/sound_api/logs/bad_files_cache.txt）'
    )
    parser.add_argument(
        '--bad-bearings',
        type=str,
        default='datasets/sound_api/logs/bad_bearings.txt',
        help='t 不连续的 bearing 列表（默认: datasets/sound_api/logs/bad_bearings.txt）'
    )
    parser.add_argument(
        '--t-policy',
        type=str,
        choices=['renumber', 'fill_gaps'],
        default='renumber',
        help='t 分配策略：renumber=强制重编号（推荐），fill_gaps=保留显式t并补齐缺口'
    )
    parser.add_argument(
        '--orig-t-warning-threshold',
        type=float,
        default=0.3,
        help='原始 t 质量告警阈值（0-1）：问题比例超过此值时写入警告（默认: 0.3 即 30%%）'
    )
    parser.add_argument(
        '--warnings-file',
        type=str,
        default='datasets/sound_api/logs/warnings.txt',
        help='原始 t 质量告警输出文件（默认: datasets/sound_api/logs/warnings.txt）'
    )
    parser.add_argument(
        '--index-file',
        type=str,
        default=None,
        help='写出 cache index.jsonl 的路径（默认: {output_dir}/index.jsonl）'
    )
    parser.add_argument(
        '--max_samples',
        type=int,
        default=None,
        help='仅处理前 N 个文件（用于快速验证，默认不限制）'
    )
    
    args = parser.parse_args()
    
    print("=" * 80)
    print("一键构建 sound_api_cache NPZ 缓存")
    print("=" * 80)
    
    # 扫描输入文件（递归）
    print(f"\n扫描输入目录:")
    print(f"  JSON: {args.json_dir}")
    print(f"  xlsx: {args.xlsx_dir}")
    
    input_files = scan_input_files_recursive(args.json_dir, args.xlsx_dir)
    
    if len(input_files) == 0:
        print("错误: 未找到任何 JSON 或 xlsx 文件")
        return
    
    print(f"找到 {len(input_files)} 个文件（优先JSON，回退xlsx）")
    if args.max_samples is not None and args.max_samples > 0:
        input_files = input_files[:args.max_samples]
        print(f"限制为前 {len(input_files)} 个文件 (--max_samples={args.max_samples})")
    
    # 构建结构化任务（解析 bearing_id 和 t，为缺失的 t 分配值）
    print(f"\n解析 bearing_id 和 t（t_policy={args.t_policy}）...")
    tasks = build_structured_tasks(input_files, t_policy=args.t_policy)
    
    print(f"构建了 {len(tasks)} 个任务")
    
    # 原始 t 质量检查（仅在 renumber 模式下）
    if args.t_policy == 'renumber':
        print(f"\n检查原始 t 质量（告警阈值: {args.orig_t_warning_threshold*100:.0f}%）...")
        orig_t_warnings = check_original_t_quality(tasks, threshold=args.orig_t_warning_threshold)
        
        if orig_t_warnings:
            # 写入警告文件
            warnings_path = Path(args.warnings_file)
            warnings_path.parent.mkdir(parents=True, exist_ok=True)
            
            with open(warnings_path, 'w', encoding='utf-8') as f:
                f.write("=" * 80 + "\n")
                f.write("原始 t 质量告警（上游数据可能有问题）\n")
                f.write("=" * 80 + "\n\n")
                f.write(f"检测时间: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"告警阈值: {args.orig_t_warning_threshold*100:.0f}%\n")
                f.write(f"发现 {len(orig_t_warnings)} 个 bearing 的原始 t 存在问题\n\n")
                
                for bearing_id, info in sorted(orig_t_warnings.items()):
                    f.write("-" * 80 + "\n")
                    f.write(f"Bearing: {bearing_id}\n")
                    f.write(f"  总文件数: {info['total_files']}\n")
                    f.write(f"  有显式 t 的文件数: {info['files_with_orig_t']}\n")
                    f.write(f"  问题比例: {info['problem_ratio']*100:.1f}%\n")
                    f.write(f"  问题类型:\n")
                    f.write(f"    - 乱序（不严格递增）: {info['issues']['disorder']} 处\n")
                    f.write(f"    - 重复值: {info['issues']['duplicate']} 个\n")
                    f.write(f"    - 跳号/不连续: {info['issues']['gaps']} 处\n")
                    f.write(f"  原始 t 样本（前20个）: {info['orig_t_values']}\n")
                    f.write("\n")
                
                f.write("=" * 80 + "\n")
                f.write("建议:\n")
                f.write("1. 检查上游 API 转换脚本是否正确标注了 t\n")
                f.write("2. 检查源数据文件名是否规范（XJTU-SY_{bearing_id}_t{t}）\n")
                f.write("3. 如果数据本身就缺失/乱序，renumber 模式会自动修正（训练不受影响）\n")
                f.write("=" * 80 + "\n")
            
            print(f"  [WARN] {len(orig_t_warnings)} bearings with orig_t issues (written to {args.warnings_file})")
            print(f"  注意：这些问题已被 renumber 自动修正，训练不受影响")
        else:
            print(f"  [OK] All bearings orig_t quality good")
    
    # 确定工作进程数
    if args.workers is None:
        workers = min(32, multiprocessing.cpu_count())
    else:
        workers = args.workers
    
    print(f"使用 {workers} 个并行进程")
    
    # 并行处理
    success_count = 0
    skip_count = 0
    fail_count = 0
    bad_files = []
    actual_t_map = defaultdict(list)  # 记录实际写入的 t
    index_records = []  # 记录 (t, orig_t, path, bearing_id) 用于对账与追溯
    
    start_time = time.time()
    
    with ProcessPoolExecutor(max_workers=workers) as executor:
        futures = {
            executor.submit(
                process_single_sample_structured,
                task,
                args.output_dir,
                not args.no_skip_existing,
                args.min_length,
                args.min_non_zero_ratio
            ): task
            for task in tasks
        }
        
        for future in as_completed(futures):
            task = futures[future]
            try:
                result = future.result()
                success = result[0]
                message = result[1]
                bad_reason = result[2]
                actual_t = result[3]
                fault_label = result[4] if len(result) > 4 else None
                if success:
                    if "跳过" in message:
                        skip_count += 1
                    else:
                        success_count += 1
                    # 记录实际写入的 t，可选带 fault_label 便于对账
                    if actual_t is not None:
                        actual_t_map[task['bearing_id']].append(actual_t)
                        rec = {
                            "t": int(actual_t),
                            "orig_t": int(task["original_t"]) if task.get("original_t") is not None else None,
                            "path": Path(task["source_path"]).name,
                            "bearing_id": str(task["bearing_id"]),
                        }
                        if fault_label is not None:
                            rec["fault_label"] = fault_label
                        index_records.append(rec)
                else:
                    fail_count += 1
                    bad_files.append({
                        'path': task['source_path'],
                        'bearing_id': task['bearing_id'],
                        't': task['t'],
                        'reason': bad_reason or message
                    })
                    print(f"失败: {task['source_path']} - {message}")
            except Exception as e:
                fail_count += 1
                bad_files.append({
                    'path': task['source_path'],
                    'bearing_id': task['bearing_id'],
                    't': task['t'],
                    'reason': f'exception: {str(e)}'
                })
                print(f"异常: {task['source_path']} - {str(e)}")
    
    elapsed_time = time.time() - start_time
    
    # 统计输出
    print("\n" + "=" * 80)
    print("处理完成")
    print("=" * 80)
    print(f"成功: {success_count} 个文件")
    print(f"跳过: {skip_count} 个文件（已存在）")
    print(f"失败: {fail_count} 个文件")
    print(f"总耗时: {elapsed_time:.2f} 秒")
    if len(tasks) > 0:
        print(f"平均速度: {len(tasks) / elapsed_time:.2f} 文件/秒")
    
    # 按失败原因分类统计
    if bad_files:
        reason_counts = defaultdict(int)
        for item in bad_files:
            reason = item['reason'].split(':')[0] if ':' in item['reason'] else item['reason']
            reason_counts[reason] += 1
        
        print("\n失败原因统计:")
        for reason, count in sorted(reason_counts.items(), key=lambda x: -x[1]):
            print(f"  {reason}: {count} 个文件")
    
    # 写入 bad_files.txt
    if bad_files:
        bad_files_path = Path(args.bad_files)
        bad_files_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(bad_files_path, 'w', encoding='utf-8') as f:
            f.write("缓存构建问题文件列表:\n")
            f.write("=" * 80 + "\n")
            for item in bad_files:
                f.write(f"文件: {item['path']}\n")
                f.write(f"bearing_id: {item['bearing_id']}\n")
                f.write(f"t: {item['t']}\n")
                f.write(f"原因: {item['reason']}\n")
                f.write("-" * 80 + "\n")
        
        print(f"\n问题文件列表已写入: {bad_files_path}")

    # 写出 index.jsonl（用于追溯与完整性对账）
    if index_records:
        index_path = Path(args.index_file) if args.index_file else (Path(args.output_dir) / "index.jsonl")
        index_path.parent.mkdir(parents=True, exist_ok=True)

        # 稳定排序：bearing_id, t
        index_records_sorted = sorted(index_records, key=lambda r: (r["bearing_id"], r["t"]))
        with open(index_path, "w", encoding="utf-8") as f:
            for rec in index_records_sorted:
                f.write(json.dumps(rec, ensure_ascii=False) + "\n")

        print(f"\ncache index 已写入: {index_path}（共 {len(index_records_sorted)} 条）")
    
    # t 连续性硬校验
    print("\n" + "=" * 80)
    print("t 连续性校验")
    print("=" * 80)
    
    bad_bearings = verify_t_continuity(args.output_dir)
    
    if bad_bearings:
        print(f"发现 {len(bad_bearings)} 个 bearing 的 t 不连续（下方仅展示前 20 个，完整列表见 {args.bad_bearings}）:")
        for idx, (bearing_id, missing_t) in enumerate(bad_bearings.items()):
            if idx >= 20:
                break
            print(f"  {bearing_id}: 缺失 t = {missing_t}")

        # 写入 bad_bearings.txt
        bad_bearings_path = Path(args.bad_bearings)
        bad_bearings_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(bad_bearings_path, 'w', encoding='utf-8') as f:
            f.write("t 不连续的 bearing 列表:\n")
            f.write("=" * 80 + "\n")
            for bearing_id, missing_t in bad_bearings.items():
                f.write(f"bearing_id: {bearing_id}\n")
                f.write(f"缺失 t: {missing_t}\n")
                f.write("-" * 80 + "\n")
        
        print(f"\nt 不连续的 bearing 列表已写入: {bad_bearings_path}")
    else:
        print("All bearings t continuous [OK]")
    
    # 检查输出目录结构
    output_path = Path(args.output_dir)
    if output_path.exists():
        bearing_dirs = [d for d in output_path.iterdir() if d.is_dir()]
        total_npz = sum(len(list(d.glob('*.npz'))) for d in bearing_dirs)
        print(f"\n输出目录结构:")
        print(f"  bearing 目录数: {len(bearing_dirs)}")
        print(f"  总 NPZ 文件数: {total_npz}")
    
    print("\n" + "=" * 80)
    print("完成！")
    print("=" * 80)


if __name__ == '__main__':
    # Windows 多进程需要这个保护
    multiprocessing.freeze_support()
    main()
