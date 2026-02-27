"""
构建 sound_api_output 目录下 xlsx 文件的索引

扫描 datasets/sound_api_output/ 下的所有 .xlsx 文件，
解析 bearing_id 和 t（时间序号），生成 index.csv。

bearing_id 解析优先级：
1. 文件名正则：XJTU-SY_(.+?)(?:_t?\d+)?\.xlsx
2. 打开 xlsx 第 1 行文本解析：XJTU-SY_(.+)

t 解析：
- 支持 _t000123 或 _000123 格式
- 若无序号，则同 bearing 按文件名排序后赋值 t=0..T-1

一致性校验：
- 每个 bearing 的 t 必须连续递增
- 不满足则写入 bad_files.txt
"""
import os
import re
import csv
import argparse
from pathlib import Path
from typing import List, Dict, Tuple, Optional
from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook


def parse_bearing_id_from_filename(filename: str) -> Optional[str]:
    """
    从文件名解析 bearing_id
    
    优先级：
    1. 正则匹配：XJTU-SY_(.+?)(?:_t?\d+)?\.xlsx
    2. 返回 None（需要从文件内容解析）
    """
    # 移除扩展名
    base_name = Path(filename).stem
    
    # 正则匹配：XJTU-SY_(.+?)(?:_t?\d+)?\.xlsx
    pattern = r'XJTU-SY_(.+?)(?:_t?\d+)?$'
    match = re.match(pattern, base_name)
    if match:
        return match.group(1)
    
    return None


def parse_bearing_id_from_xlsx_first_row(xlsx_path: str) -> Optional[str]:
    """
    从 xlsx 文件第 1 行文本解析 bearing_id
    
    格式：XJTU-SY_(.+)
    """
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        
        # 读取第 1 行第 1 列
        first_cell = ws.cell(row=1, column=1)
        if first_cell.value:
            text = str(first_cell.value)
            pattern = r'XJTU-SY_(.+)'
            match = re.search(pattern, text)
            if match:
                wb.close()
                return match.group(1)
        
        wb.close()
    except Exception as e:
        print(f"Warning: 无法读取 {xlsx_path} 第 1 行: {e}")
    
    return None


def parse_t_from_filename(filename: str) -> Optional[int]:
    """
    从文件名解析 t（时间序号）
    
    支持格式：
    - _t000123
    - _000123
    """
    base_name = Path(filename).stem
    
    # 尝试 _t000123 格式
    pattern1 = r'_t(\d+)$'
    match = re.search(pattern1, base_name)
    if match:
        return int(match.group(1))
    
    # 尝试 _000123 格式（6位数字）
    pattern2 = r'_(\d{6})$'
    match = re.search(pattern2, base_name)
    if match:
        return int(match.group(1))
    
    return None


def scan_xlsx_files(data_dir: str) -> List[Dict[str, str]]:
    """
    扫描目录下的所有 xlsx 文件，解析 bearing_id 和 t
    
    Returns:
        List[Dict]: 每个元素包含 path, bearing_id, t
    """
    data_path = Path(data_dir)
    if not data_path.exists():
        raise FileNotFoundError(f"数据目录不存在: {data_dir}")
    
    records = []
    bad_files = []
    
    xlsx_files = list(data_path.glob('*.xlsx'))
    print(f"找到 {len(xlsx_files)} 个 xlsx 文件")
    
    for xlsx_file in xlsx_files:
        path_str = str(xlsx_file)
        
        # 解析 bearing_id
        bearing_id = parse_bearing_id_from_filename(xlsx_file.name)
        if bearing_id is None:
            bearing_id = parse_bearing_id_from_xlsx_first_row(path_str)
        
        if bearing_id is None:
            bad_files.append({
                'path': path_str,
                'reason': '无法解析 bearing_id'
            })
            continue
        
        # 解析 t
        t = parse_t_from_filename(xlsx_file.name)
        
        records.append({
            'path': path_str,
            'bearing_id': bearing_id,
            't': t  # 可能为 None
        })
    
    # 对于没有 t 的文件，按 bearing_id 分组，按文件名排序后赋值 t=0..T-1
    bearing_groups = defaultdict(list)
    for record in records:
        if record['t'] is None:
            bearing_groups[record['bearing_id']].append(record)
    
    for bearing_id, group in bearing_groups.items():
        # 按文件名排序
        group.sort(key=lambda x: x['path'])
        for idx, record in enumerate(group):
            record['t'] = idx
    
    # 转换 t 为整数
    for record in records:
        if record['t'] is not None:
            record['t'] = int(record['t'])
    
    return records, bad_files


def validate_t_continuity(records: List[Dict]) -> Tuple[List[Dict], List[Dict]]:
    """
    校验每个 bearing 的 t 是否连续递增
    
    Returns:
        valid_records: 有效的记录
        bad_records: 无效的记录（t 不连续或重复）
    """
    bearing_groups = defaultdict(list)
    for record in records:
        bearing_groups[record['bearing_id']].append(record)
    
    valid_records = []
    bad_records = []
    
    for bearing_id, group in bearing_groups.items():
        # 按 t 排序
        group.sort(key=lambda x: x['t'])
        
        # 检查是否连续
        t_values = [r['t'] for r in group]
        expected_t = list(range(len(t_values)))
        
        if t_values != expected_t:
            # t 不连续或重复
            bad_records.extend(group)
            print(f"Warning: bearing_id={bearing_id} 的 t 不连续: {t_values}")
        else:
            valid_records.extend(group)
    
    return valid_records, bad_records


def write_index_csv(records: List[Dict], output_path: str):
    """写入 index.csv"""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=['path', 'bearing_id', 't'])
        writer.writeheader()
        for record in records:
            writer.writerow({
                'path': record['path'],
                'bearing_id': record['bearing_id'],
                't': record['t']
            })
    
    print(f"索引已写入: {output_path}")


def write_bad_files(bad_files: List[Dict], output_path: str):
    """写入 bad_files.txt"""
    if not bad_files:
        return
    
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("问题文件列表:\n")
        f.write("=" * 80 + "\n")
        for item in bad_files:
            f.write(f"文件: {item['path']}\n")
            f.write(f"原因: {item['reason']}\n")
            f.write("-" * 80 + "\n")
    
    print(f"问题文件列表已写入: {output_path}")


def main():
    parser = argparse.ArgumentParser(description='构建 sound_api_output 索引')
    parser.add_argument(
        '--data_dir',
        type=str,
        default='datasets/sound_api_output',
        help='xlsx 文件目录（默认: datasets/sound_api_output）'
    )
    parser.add_argument(
        '--output',
        type=str,
        default='datasets/sound_api_output/index.csv',
        help='输出索引文件路径（默认: datasets/sound_api_output/index.csv）'
    )
    parser.add_argument(
        '--bad_files',
        type=str,
        default='datasets/sound_api_output/bad_files.txt',
        help='问题文件列表路径（默认: datasets/sound_api_output/bad_files.txt）'
    )
    parser.add_argument(
        '--strict',
        action='store_true',
        help='严格模式：如果存在 t 不连续的情况，报错停止'
    )
    
    args = parser.parse_args()
    
    print("=" * 80)
    print("构建 sound_api_output 索引")
    print("=" * 80)
    
    # 扫描文件
    print(f"\n扫描目录: {args.data_dir}")
    records, bad_files_parse = scan_xlsx_files(args.data_dir)
    
    if len(records) == 0:
        print("错误: 未找到任何有效的 xlsx 文件")
        return
    
    print(f"成功解析 {len(records)} 个文件")
    if bad_files_parse:
        print(f"警告: {len(bad_files_parse)} 个文件无法解析 bearing_id")
    
    # 校验 t 连续性
    print("\n校验 t 连续性...")
    valid_records, bad_records = validate_t_continuity(records)
    
    if bad_records:
        print(f"警告: {len(bad_records)} 个文件的 t 不连续或重复")
        bad_files_parse.extend([
            {
                'path': r['path'],
                'reason': f"t 不连续或重复 (bearing_id={r['bearing_id']}, t={r['t']})"
            }
            for r in bad_records
        ])
        
        if args.strict:
            print("错误: 严格模式启用，停止处理")
            write_bad_files(bad_files_parse, args.bad_files)
            return
    
    # 统计信息
    bearing_count = len(set(r['bearing_id'] for r in valid_records))
    total_samples = len(valid_records)
    
    print(f"\n统计信息:")
    print(f"  有效 bearing 数: {bearing_count}")
    print(f"  总样本数: {total_samples}")
    print(f"  平均每个 bearing 样本数: {total_samples / bearing_count:.1f}")
    
    # 写入索引
    write_index_csv(valid_records, args.output)
    
    # 写入问题文件列表
    if bad_files_parse:
        write_bad_files(bad_files_parse, args.bad_files)
        print(f"\n警告: 发现 {len(bad_files_parse)} 个问题文件，已写入 {args.bad_files}")
    else:
        print("\n所有文件解析成功！")


if __name__ == '__main__':
    main()
