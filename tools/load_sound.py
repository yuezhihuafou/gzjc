"""
声音能量曲线数据加载器
从 xlsx 文件中加载声音密度和能量数据

支持多sheet结构：
- 每个xlsx可包含多个sheet
- 每个sheet对应一个原始样本
- Sheet名称格式："{filename}.wav"
"""
import os
import pandas as pd
import numpy as np
from pathlib import Path

class SoundDataLoader:
    """声音能量曲线数据加载器 - 支持多sheet结构 (改进版)"""
    
    def __init__(self, sound_data_dir='声音能量曲线数据', lazy_load=True):
        """
        Args:
            sound_data_dir: 声音数据目录路径
            lazy_load: 是否延迟加载（True=只在需要时读取文件，False=初始化时读取所有文件）
                      建议设置为True以避免初始化时消耗大量内存
        """
        self.sound_data_dir = sound_data_dir
        self.lazy_load = lazy_load
        self._sheet_to_file = None  # 延迟初始化
        self._file_mapping = None    # 延迟初始化
        
        # 只有在非延迟模式下才立即构建映射
        if not lazy_load:
            self._sheet_to_file = self._build_sheet_mapping()
            self._file_mapping = self._build_file_mapping()
    
    @property
    def sheet_to_file(self):
        """延迟加载：只在首次访问时构建映射"""
        if self._sheet_to_file is None:
            self._sheet_to_file = self._build_sheet_mapping()
        return self._sheet_to_file
    
    @property
    def file_mapping(self):
        """延迟加载：只在首次访问时构建映射"""
        if self._file_mapping is None:
            self._file_mapping = self._build_file_mapping()
        return self._file_mapping
        
    def _build_sheet_mapping(self):
        """
        构建sheet到xlsx文件的映射
        返回: {
            '97_Normal_0': ('path/to/normal.xlsx', '97_Normal_0.wav'),
            '234_0': ('path/to/04.xlsx', '234_0.wav'),
            ...
        }
        """
        mapping = {}
        sound_dir = Path(self.sound_data_dir)
        
        if not sound_dir.exists():
            print(f"Warning: Sound data directory {sound_dir} not found")
            return mapping
        
        # 只读取文件列表，不读取文件内容（节省内存）
        for xlsx_file in sound_dir.glob('*.xlsx'):
            try:
                # 使用 openpyxl 只读取工作表名称，不加载数据内容
                # 这样可以避免将整个文件加载到内存
                from openpyxl import load_workbook
                wb = load_workbook(xlsx_file, read_only=True, data_only=False)
                for sheet_name in wb.sheetnames:
                    # sheet_name 格式: "97_Normal_0.wav" 或 "234_0.wav"
                    base_name = sheet_name.replace('.wav', '')
                    mapping[base_name] = (str(xlsx_file), sheet_name)
                wb.close()
            except Exception as e:
                # 如果 openpyxl 失败，回退到 pandas（但会消耗更多内存）
                try:
                    xls = pd.ExcelFile(xlsx_file)
                    for sheet_name in xls.sheet_names:
                        base_name = sheet_name.replace('.wav', '')
                        mapping[base_name] = (str(xlsx_file), sheet_name)
                except Exception as e2:
                    print(f"Warning: Failed to read {xlsx_file}: {e2}")
        
        return mapping
    
    def _build_file_mapping(self):
        """构建兼容旧版本的映射"""
        mapping = {}
        for base_name, (xlsx_path, sheet_name) in self.sheet_to_file.items():
            mapping[base_name] = xlsx_path
        return mapping
    
    def load_sound_curves(self, filename_base):
        """
        加载指定文件的声音能量曲线
        
        Args:
            filename_base: 基础文件名，如 "97_Normal_0" 或 "234_0"
            
        Returns:
            dict: {
                'frequency': np.array,  # 频率 (Hz)
                'volume': np.array,     # 音量
                'density': np.array     # 密度
            }
            如果文件不存在返回 None
        """
        # 尝试多种命名格式
        possible_names = [
            filename_base,
            filename_base.replace('.mat', ''),
            filename_base.split('/')[-1].replace('.mat', '')
        ]
        
        xlsx_path = None
        sheet_name = None
        
        # 使用属性访问，触发延迟加载
        for name in possible_names:
            if name in self.sheet_to_file:
                xlsx_path, sheet_name = self.sheet_to_file[name]
                break
        
        if xlsx_path is None:
            return None
        
        try:
            # 读取指定sheet的数据（跳过前2行：文件名和列标题）
            df = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=None, skiprows=2)
            
            return {
                'frequency': df.iloc[:, 0].values.astype(np.float32),  # 第一列：频率
                'volume': df.iloc[:, 1].values.astype(np.float32),     # 第二列：音量
                'density': df.iloc[:, 2].values.astype(np.float32)     # 第三列：密度
            }
        except Exception as e:
            print(f"Error loading {xlsx_path}[{sheet_name}]: {e}")
            return None
    
    def get_available_files(self):
        """返回所有可用的声音数据文件列表"""
        # 使用属性访问，触发延迟加载
        return list(self.sheet_to_file.keys())
    
    def get_statistics(self):
        """返回数据集统计信息"""
        # 使用属性访问，触发延迟加载
        return {
            'total_files': len(self.sheet_to_file),
            'available_files': list(self.sheet_to_file.keys())
        }


def test_loader():
    """测试加载器功能"""
    loader = SoundDataLoader()
    
    print("=== Sound Data Loader Statistics ===")
    stats = loader.get_statistics()
    print(f"Total files: {stats['total_files']}")
    print(f"\nFirst 10 available files:")
    for f in stats['available_files']:
        print(f"  - {f}")
    
    # 测试加载一个文件
    print("\n=== Testing load_sound_curves ===")
    test_file = stats['available_files'][0] if stats['available_files'] else None
    
    if test_file:
        curves = loader.load_sound_curves(test_file)
        if curves:
            print(f"\nLoaded curves for: {test_file}")
            print(f"  Frequency range: {curves['frequency'].min():.2f} - {curves['frequency'].max():.2f} Hz")
            print(f"  Volume range: {curves['volume'].min():.2f} - {curves['volume'].max():.2f}")
            print(f"  Density range: {curves['density'].min():.2f} - {curves['density'].max():.2f}")
            print(f"  Number of points: {len(curves['frequency'])}")
        else:
            print(f"Failed to load curves for {test_file}")
    
    return loader


if __name__ == '__main__':
    test_loader()
